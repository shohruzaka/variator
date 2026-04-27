"""Excel (xlsx) formatda javoblar kalitini eksport qilish (8-bosqich)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from src.models import Variant


def export_answers_to_xlsx(
    variants: list[Variant],
    output_path: str | Path,
    subject_name: str = "",
    assessment_type: str = "",
) -> Path:
    """Barcha variantlarning javoblarini bitta Excel fayliga yozadi.

    Jadval tuzilishi:
    Qatorlar: Variant raqamlari
    Ustunlar: Variant No, 1-savol, 2-savol, ...

    Args:
        variants: Generatsiya qilingan test variantlari ro'yxati.
        output_path: Saqlanadigan Excel faylining to'liq manzili.
        subject_name: Fan nomi (bo'sh bo'lsa yozilmaydi).
        assessment_type: Nazorat turi (bo'sh bo'lsa yozilmaydi).

    Returns:
        Yaratilgan faylning manzili.
    """
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Javoblar kaliti"

    if not variants:
        wb.save(out_path)
        return out_path

    # Maksimal savollar sonini aniqlash (ustunlar uchun)
    max_questions = max(len(v.questions) for v in variants)
    total_cols = max_questions + 1  # "Variant" ustunini hisobga olib

    label_font = Font(bold=True)

    # Fan nomi va nazorat turi qatorlari (jadvaldan oldin)
    if subject_name:
        ws.append(["Fan:", subject_name])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=1).font = label_font
        if total_cols > 2:
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=total_cols)

    if assessment_type:
        ws.append(["Nazorat turi:", assessment_type])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=1).font = label_font
        if total_cols > 2:
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=total_cols)

    # 1. Sarlavha qatorini yozish
    header = ["Variant"] + [str(i + 1) for i in range(max_questions)]
    ws.append(header)
    header_row = ws.max_row

    # Sarlavha dizayni (Qalin harf va markazlashtirilgan)
    header_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[header_row]:
        cell.font = label_font
        cell.alignment = header_align

    # 2. Har bir variant javoblarini yozish
    for variant in variants:
        row_data = [f"{variant.number}-variant"] + variant.answer_key
        ws.append(row_data)

    # Hujayralarni markazlashtirish (Variant nomidan tashqari)
    for row in ws.iter_rows(min_row=header_row + 1, max_col=total_cols):
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center")

    # Ustunlar kengligini to'g'rilash
    ws.column_dimensions["A"].width = 15
    for i in range(2, max_questions + 2):
        col_letter = ws.cell(row=header_row, column=i).column_letter
        ws.column_dimensions[col_letter].width = 5

    wb.save(out_path)
    return out_path
