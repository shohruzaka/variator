"""Exporter (Excel) uchun unit testlar (8-bosqich)."""

from pathlib import Path

import openpyxl
import pytest

from src.exporter_xlsx import export_answers_to_xlsx
from src.models import Option, Question, Variant


def _make_sample_variants() -> list[Variant]:
    """Test uchun 2 ta variant yaratadi."""
    q1 = Question(1, "S1", [Option("A", "a"), Option("B", "b"), Option("C", "c", is_correct=True), Option("D", "d")])
    q2 = Question(2, "S2", [Option("A", "a"), Option("B", "b"), Option("C", "c"), Option("D", "d", is_correct=True)])
    v1 = Variant(number=1, seed=1, questions=[q1, q2])
    
    q3 = Question(1, "S2", [Option("A", "a"), Option("B", "b", is_correct=True), Option("C", "c"), Option("D", "d")])
    q4 = Question(2, "S1", [Option("A", "a", is_correct=True), Option("B", "b"), Option("C", "c"), Option("D", "d")])
    v2 = Variant(number=2, seed=2, questions=[q3, q4])
    
    return [v1, v2]


def test_export_answers_to_xlsx(tmp_path: Path):
    """Excel fayl to'g'ri yaratilishi va ma'lumotlar kiritilishini tekshirish."""
    variants = _make_sample_variants()
    out_file = tmp_path / "javoblar.xlsx"
    
    result_path = export_answers_to_xlsx(variants, out_file)
    
    assert result_path.exists()
    
    # Faylni o'qib tekshiramiz
    wb = openpyxl.load_workbook(result_path)
    ws = wb.active
    
    assert ws.title == "Javoblar kaliti"
    
    # Qatorlarni o'qiymiz
    rows = list(ws.iter_rows(values_only=True))
    
    # 1. Sarlavha: Variant, 1, 2
    assert rows[0] == ("Variant", "1", "2")
    
    # 2. 1-Variant javoblari (q1=C, q2=D)
    assert rows[1] == ("1-Variant", "C", "D")
    
    # 3. 2-Variant javoblari (q3=B, q4=A)
    assert rows[2] == ("2-Variant", "B", "A")


def test_export_answers_empty_variants_xlsx(tmp_path: Path):
    """Bo'sh ro'yxat bo'lganda ham Excel fayl yaratilishi kerak."""
    out_file = tmp_path / "javoblar_bosh.xlsx"
    result_path = export_answers_to_xlsx([], out_file)
    
    assert result_path.exists()
    
    wb = openpyxl.load_workbook(result_path)
    ws = wb.active
    
    # Faylda faqat bitta bo'sh jadval bo'lishi kerak
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 0 or (len(rows) == 1 and not any(rows[0]))
